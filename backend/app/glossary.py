from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Dict, Iterable, List, Optional

from openpyxl import load_workbook

from .models import GlossarySummary


@dataclass(frozen=True)
class GlossaryTerm:
    source: str
    target: str
    category: str
    fixed: bool
    confidence: str
    note: str


class Glossary:
    def __init__(self, terms: Iterable[GlossaryTerm], source: Optional[str] = None):
        self.terms: List[GlossaryTerm] = list(terms)
        self.source = source
        self.by_source: Dict[str, GlossaryTerm] = {
            term.source: term for term in self.terms if term.source
        }
        self.by_normalized: Dict[str, GlossaryTerm] = {
            self.normalize(term.source): term for term in self.terms if term.source
        }
        self._longest_first = sorted(
            self.terms, key=lambda term: len(term.source), reverse=True
        )

    @classmethod
    def empty(cls) -> "Glossary":
        return cls([])

    @classmethod
    def from_xlsx(cls, path: Path) -> "Glossary":
        workbook = load_workbook(path, read_only=True, data_only=True)
        if "术语库" not in workbook.sheetnames:
            raise ValueError("术语表缺少“术语库”工作表")
        sheet = workbook["术语库"]
        headers = [str(cell.value or "").strip() for cell in sheet[1]]
        index = {name: position for position, name in enumerate(headers)}
        required = ["日文原文", "中文标准译法", "是否必须固定", "置信度"]
        missing = [name for name in required if name not in index]
        if missing:
            raise ValueError(f"术语表缺少字段：{', '.join(missing)}")

        def value(row, name: str) -> str:
            position = index.get(name)
            if position is None or position >= len(row):
                return ""
            return str(row[position] or "").strip()

        terms = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            source = value(row, "日文原文")
            target = value(row, "中文标准译法")
            if not source or not target:
                continue
            confidence = value(row, "置信度")
            terms.append(
                GlossaryTerm(
                    source=source,
                    target=target,
                    category=value(row, "类别"),
                    fixed=value(row, "是否必须固定") == "是" and confidence == "高",
                    confidence=confidence,
                    note=value(row, "备注/使用规则"),
                )
            )
        return cls(terms, source=path.name)

    def exact(self, text: str) -> Optional[GlossaryTerm]:
        stripped = text.strip()
        return self.by_source.get(stripped) or self.by_normalized.get(self.normalize(stripped))

    def matches(self, text: str) -> List[GlossaryTerm]:
        normalized = self.normalize(text)
        return [
            term for term in self._longest_first
            if term.source in text or self.normalize(term.source) in normalized
        ]

    @staticmethod
    def normalize(text: str) -> str:
        return re.sub(r"\s+", "", text).replace("／", "/").replace("～", "〜")

    def summary(self) -> GlossarySummary:
        fixed = sum(1 for term in self.terms if term.fixed)
        return GlossarySummary(
            total=len(self.terms),
            fixed=fixed,
            review=len(self.terms) - fixed,
            source=self.source,
        )
